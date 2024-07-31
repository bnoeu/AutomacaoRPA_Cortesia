import openpyxl
import pandas as pd
import time
import os
from onedrivedownloader import download as one_download

#* Funções

#! Essa função só deve ser executada não houver linhas com a coluna "Status" vazias no "producao"
def baixa_planilha():
    try:
        os.remove("db_alltrips/db_alltrips.xlsx")
    except FileNotFoundError:
        print('Arquivo db_alltrips.xlsx já excluido.')
    except PermissionError:
        exit(print('--- Planilha aberta, fechar a planilha'))

    ln = "https://cortesiaconcreto-my.sharepoint.com/:x:/g/personal/bi_cortesiaconcreto_com_br/EW_8FZwWFYVAol4MpV1GglkBJEaJaDx6cfuClnesIu60Ng?e=pveECF"
    one_download(ln, filename="db_alltrips")

def verifica_status_vazio(ultima_linha = 0):
    linha_atual = []
    conta_linha = 1
    for rows in producao.iter_rows(min_row= 2, max_row = ultima_linha):
        conta_linha += 1
        if rows[6].value is None:
            for item in range(0, 9):
                linha_atual.append(rows[item].value)
            print(F'Existem linhas vazias:  {linha_atual}')
            break

    print(F'Ultima linha com campo status vazio: {conta_linha}')
    return linha_atual, conta_linha    
#* --------------------------------------------------- PROGRAMA PRINCIPAL ---------------------------------------------------

# Reorganiza o arquivo.
def reorganiza_arquivo(nome = ""):
    dataframe = pd.read_excel('db_alltrips/' + nome, date_format= "dd/mm/yyyy") #Insere o arquivo no dataframe
    dataframe['D. Insercao'] = pd.to_datetime(dataframe["D. Insercao"], errors= 'coerce', format= '%d/%m/%Y') # Converte a coluna para o tipo "datetime"
    #! Vou precisar deletar as linhas sem a data de insercao?!
    #dataframe = dataframe.dropna(subset=["D. Insercao"])
    dataframe = dataframe.sort_values(by = "D. Insercao") # Reordena o arquivo, com base na coluna "D. Insercao"
    dataframe.to_excel('db_alltrips/ordenado_' + nome, index= False)

reorganiza_arquivo(nome = "db_alltrips.xlsx")
exit()
# Carrega os arquivos
book = openpyxl.load_workbook("db_alltrips/db_alltrips.xlsx")
try:
    book2 = openpyxl.load_workbook("db_alltrips/producao.xlsx")
except FileNotFoundError:
    book.save("db_alltrips/producao.xlsx")

# Seleciona uma pagina
db_alltrips =  book['db_alltrips']
producao = book2['db_alltrips']

# Retorna o tamanho maximo da planilha
ultima_linha_db = db_alltrips.max_row
ultima_linha_producao = producao.max_row
print(F'--- Ultima linha: DB: {ultima_linha_db}, Produção: {ultima_linha_producao}')
dados_atuais = verifica_status_vazio(ultima_linha_producao)
print(dados_atuais[0])


#* Coleta os novos dados em db_alltrips e move para a producao
linha_atual = []
quantidade_novas_linhas = ultima_linha_db - ultima_linha_producao
print(F"--- Novas linhas detectadas! {quantidade_novas_linhas}")
if quantidade_novas_linhas > 0:
    #1. Descobrir a ultima linha da producao
    ultima_linha_producao = producao.max_row
    #2. Com base na ultima linha da produção, coletar a proxima linha do alltrips
    for rows in db_alltrips.iter_rows(min_row= ultima_linha_producao + 1, max_row =  ultima_linha_db):
        if rows[6].value is None:
            for item in range(0, 9):
                linha_atual.append(rows[item].value)
            print(F'Dados coletado {linha_atual}')
            #Insere a linha na producao
            producao.append(linha_atual)
            linha_atual = []
        
    book2.save("db_alltrips/producao.xlsx")

exit()



for rows in producao.iter_rows(min_row= ultima_linha_producao, max_row = ultima_linha_producao):
    rows[6].value = "Teste_bruno"


book2.save("db_alltrips/producao.xlsx")
